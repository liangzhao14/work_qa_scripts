import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_questions_from_txt(file_path):
    data_structure = {}  # {"章节": {"节": [{"问题": "", "答案": ""}, ...]}}
    current_chapter = None
    current_section = None
    current_question = None
    current_answer_lines = []  # 用于存储多行答案内容

    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    i = 0
    n = len(lines)
    while i < n:
        line = lines[i].strip()
        if not line:
            i += 1
            # 如果遇到空行，且我们已经有了问题与答案行，则尝试保存
            if current_question and current_answer_lines:
                answer = '\n'.join(current_answer_lines).strip()
                if current_chapter and current_section and current_question:
                    if current_chapter not in data_structure:
                        data_structure[current_chapter] = {}
                    if current_section not in data_structure[current_chapter]:
                        data_structure[current_chapter][current_section] = []
                    data_structure[current_chapter][current_section].append({
                        "问题": current_question,
                        "答案": answer
                    })
                # 重置
                current_question = None
                current_answer_lines = []
            continue

        if line.startswith('【章节】'):
            current_chapter = line.replace('【章节】', '').strip()
            if current_chapter not in data_structure:
                data_structure[current_chapter] = {}
            i += 1

        elif line.startswith('【节】'):
            current_section = line.replace('【节】', '').strip()
            if current_section not in data_structure[current_chapter]:
                data_structure[current_chapter][current_section] = []
            i += 1

        elif line.startswith('【问题】'):
            current_question = line.replace('【问题】', '').strip()
            current_answer_lines = []  # 重置答案行
            i += 1

        elif line.startswith('【答案】'):
            # 先读取【答案】这一行
            answer_part = line.replace('【答案】', '').strip()
            current_answer_lines.append(answer_part)
            i += 1

            # 继续读取后续行，直到遇到空行、或下一个【问题】
            while i < n:
                next_line = lines[i].strip()
                if not next_line:
                    # 遇到空行，结束答案读取
                    break
                if next_line.startswith('【章节】') or next_line.startswith('【节】') or next_line.startswith('【问题】'):
                    # 遇到新的结构，结束答案读取
                    break
                # 否则认为是答案的延续
                current_answer_lines.append(next_line)
                i += 1

            # 所有答案行收集完毕，合并并保存
            if current_question and current_answer_lines:
                answer = '\n'.join(current_answer_lines).strip()
                if current_chapter and current_section:
                    if current_chapter not in data_structure:
                        data_structure[current_chapter] = {}
                    if current_section not in data_structure[current_chapter]:
                        data_structure[current_chapter][current_section] = []
                    data_structure[current_chapter][current_section].append({
                        "问题": current_question,
                        "答案": answer
                    })
                # 重置
                current_question = None
                current_answer_lines = []

        else:
            i += 1

    return data_structure


def create_excel_from_structure(data_structure, output_file="电力安全问答库.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "问答库"

    headers = ["问答库名称（章节）", "类别名称（节）", "标准问题", "答案"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col, value=headers[col - 1])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row_num = 2
    for chapter, sections in data_structure.items():
        for section, questions in sections.items():
            for qa in questions:
                ws.append([chapter, section, qa["问题"], qa["答案"]])
                row_num += 1

    column_widths = [25, 20, 60, 80]
    for i, width in enumerate(column_widths, 1):
        if i <= 4:
            ws.column_dimensions[chr(64 + i)].width = width

    wb.save(output_file)
    print(f"✅ Excel 文件已生成：{output_file}")
    print(f"📊 共生成 {row_num - 1} 条问答")

if __name__ == "__main__":
    input_txt = input("请输入您的问答文本文件路径（如：questions.txt）: ").strip()
    if not input_txt:
        input_txt = "questions.txt"

    output_excel = input("请输入输出的 Excel 文件名（直接回车默认为 电力安全问答库.xlsx）: ").strip()
    if not output_excel:
        output_excel = "电力安全问答库.xlsx"
    elif not output_excel.endswith(".xlsx"):
        output_excel += ".xlsx"

    data = parse_questions_from_txt(input_txt)
    create_excel_from_structure(data, output_excel)