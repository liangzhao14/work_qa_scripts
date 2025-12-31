import requests
import json
import pandas as pd
import time
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
# from PIL import Image
# from xlsxwriter.utility import xl_rowcol_to_cell
# import xlsxwriter
import io
import numpy as np


class AskRequsets():
    def __init__(self, localhost):
        self.df = pd.DataFrame(columns=['docId', 'mainVenifyContent', 'question', 'answer'])

        self.localhost = localhost
        # self.token = self.login_system("DTadminpro", "DT@admin")
        self.token = self.login_system("minzuo111", "1234qwer")
        self.outputData = []
        self.file_path = r'C:/Users/admin/PycharmProjects/PythonProject/AskRequest/result3.xlsx'
        self.columns = ['docId', 'mainVenifyContent', 'queCstion', 'answer', 'LLMresult', 'paragraph1', 'paragraph2',
                        'paragraph3', 'paragraph4', 'paragraph5']

    def login_system(self, username, password):
        url = f'{self.localhost}/digital-artisan-user/login/loginWithPassword'

        body = {"username": username, "password": password}
        headers = {"Content-Type": "application/json"}
        response = requests.post(url=url, headers=headers, json=body, verify=False)
        return response.json()['data']['token']

    def make_request(self, question, answer, assistantId, spaceId, docId, mainVenifyContent):
        global data_dict
        print(f"开始提问大模型：{question}")
        # print(question,answer,assistantId,spaceId,docId,mainVenifyContent,"444444444")
        T = 1
        # 问3次
        while T <= 3:
            try:
                # print(T)
                url = f"{self.localhost}/digital-artisan-chat2/botChat/stream"
                body = {"question": question, "spaceIds": spaceId, "docIds": [], "assistantId": assistantId,
                        "conversationId": "CHAT_1736144105620", "requestId": "584796f7-6a7a-42cd-81f7-cb5a8bdd6541",
                        "image": "", "selectModelIds": []}
                headers = {"Content-Type": "application/json", "Artisan-Token": self.token}
                # print(body,headers)
                response = requests.post(url=url,
                                         headers=headers, json=body, stream=True, verify=False)
                response.encoding = 'utf-8'
                print(response.text,"abhhhhhhbbggggggggggggggggggggg")
                result = ""
                recall = []
                recallparagraph = []
                paragraph1 = ""
                paragraph2 = ""
                paragraph3 = ""
                paragraph4 = ""
                paragraph5 = ""
                # and data_dict["status"] == 1
                for line in response.iter_lines(decode_unicode=True):
                    if line:
                        # print(line)
                        # 取流scream
                        data_string = line.split("data:", 1)[1]
                        # 字符串转json
                        data_dict = json.loads(data_string)
                        print(data_dict, "666666666666666666666666666666666666")

                        if 'dataMessage' in data_dict and 'sql' in data_dict["dataMessage"] and 'rows' in \
                                data_dict["dataMessage"]["dataframe"]:
                            result = result + data_dict["answerText"]

                            paragraph1 = data_dict["dataMessage"]["sql"]
                            paragraph2 = data_dict["dataMessage"]["dataframe"]["rows"][0]
                            print(result, "PPPPPPPPPP")
                            print(result, "MMMMMMMM")
                            print(result, "NNNN")


                    else:
                        pass
                break
            except:
                T += 1
                time.sleep(2)
        print(result)
        # data
        data = ''
        print(f'question:{question}\n'
              f'LLM result:{result}\n'
              f'paragraph1:{data_dict["dataMessage"]["sql"]}\n'
              f'paragraph2:{data_dict["dataMessage"]["dataframe"]["rows"][0]}\n'
              f'paragraph3:{paragraph3}\n'
              f'paragraph4:{paragraph4}\n'
              f'paragraph5:{paragraph5}\n')
        # print(paragraph1)
        # #id编号
        # createQaNumber = int(time.time() * 1000)
        item = [docId, mainVenifyContent, question, answer, result + data_dict["answerText"], data_dict["dataMessage"]["sql"], data_dict["dataMessage"]["dataframe"]["rows"][0], paragraph3, paragraph4,
                paragraph5]
        print(item)
        self.outputData.append(item)
        print(result)
        return result

    # 新增图片的写入
    def create_excel(self):
        print("开始输出excel...")
        # 检查文件是否存在
        if os.path.exists(self.file_path):
            # 如果文件存在,则以追加模式打开
            writer = pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            # 读取原有数据
            existing_data = pd.read_excel(self.file_path, sheet_name='Sheet1')
            # 合并新旧数据
            combined_data = pd.concat([existing_data, pd.DataFrame(self.outputData, columns=self.columns)],
                                      ignore_index=True)
            # 将合并后的数据写入Excel
            combined_data.to_excel(writer, sheet_name='Sheet1', index=False, header=True)
        else:
            # 如果文件不存在,则创建一个新的
            writer = pd.ExcelWriter(self.file_path, engine='openpyxl')
            # 创建DataFrame并写入Excel
            df = pd.DataFrame(data=self.outputData, columns=self.columns)
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=True)
        writer.close()
        print("Excel已输出.")
    # def writeAnswerImage(self):
    #    wb = load_workbook(r'C:\Users\fqidong\Desktop\AskRequest\result2.xlsx')
    #    ws = wb[wb.sheetnames[0]]
    #  # 创建
    #
    #    target_wb.save(self.file_path)
    #    target_wb.close()
    #    print("Excel已输出.")#一个新的Excel工作簿和工作表
    #    target_wb = load_workbook(self.file_path)
    #    if target_wb.sheetnames:
    #     target_ws = target_wb.active
    #    else:
    #     target_ws = target_wb.create_sheet()
    #    for i, image in enumerate(ws._images):
    #       anchor = image.anchor._from
    #       row = anchor.row
    #       rows = row+1
    #       image.width, image.height = 20, 20
    #       #添加图片到目标工作表
    #       target_ws.add_image(image, f'D{rows}')
    #


if __name__ == '__main__':
    a = AskRequsets(localhost="http://10.170.249.86:8800")
    b = a.make_request(question="甘肃龙源5月6日限出力是多少", assistantId=978, answer="", spaceId="", docId="",
                       mainVenifyContent="")
    a.create_excel()